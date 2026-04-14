from app.schemas.schemas import XlsxBase64Request
import json
import os
import base64
import xml.etree.ElementTree as ET
from fastapi import HTTPException, UploadFile
import re
from io import BytesIO
from typing import List
from openpyxl import load_workbook




POT_DO_DATOTEKE = "podatki/uporabniki.json"

def shrani_userja(user):
    os.makedirs("podatki", exist_ok=True)

    nov_uporabnik = {
        "ime": user.ime,
        "telefonska_stevilka": user.telefonska_stevilka,
        "datum_rojstva": str(user.datum_rojstva)
    }

    if os.path.exists(POT_DO_DATOTEKE):
        with open(POT_DO_DATOTEKE, "r", encoding="utf-8") as file:
            try:
                users = json.load(file)
            except json.JSONDecodeError:
                users = []
    else:
        users = []

    users.append(nov_uporabnik)

    with open(POT_DO_DATOTEKE, "w", encoding="utf-8") as file:
        json.dump(users, file, ensure_ascii=False, indent=4)

    return {
        "message": f"Uporabnik {user.ime} je shranjen v sistem!"
    }



def vrni_vse_userje():
    if os.path.exists(POT_DO_DATOTEKE):
        with open(POT_DO_DATOTEKE, "r", encoding="utf-8") as file:
            try:
                users = json.load(file)
            except json.JSONDecodeError:
                users = []
    else:
        users = []

    return users


async def obdelaj_xml_datoteke(payload):
    files_base64 = payload.files
    telefonske_stevilke = []

    for i, file_base64 in enumerate(files_base64):
        try:
            if "," in file_base64:
                file_base64 = file_base64.split(",", 1)[1]

            vsebina = base64.b64decode(file_base64)
            xml_text = vsebina.decode("utf-8")
            root = ET.fromstring(xml_text)

            stevilka = root.findtext(".//telefonska_stevilka")
            if stevilka:
                telefonske_stevilke.append(stevilka)

        except ET.ParseError:
            raise HTTPException(
                status_code=400,
                detail=f"Datoteka na indeksu {i} vsebuje neveljaven XML"
            )
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=400,
                detail=f"Datoteka na indeksu {i} ni veljaven UTF-8 XML"
            )
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Napaka pri obdelavi datoteke na indeksu {i}: {str(e)}"
            )

    os.makedirs("podatki", exist_ok=True)
    pot_do_datoteke = "podatki/telefonske_cifre.json"

    obstojece_stevilke = []
    if os.path.exists(pot_do_datoteke):
        try:
            with open(pot_do_datoteke, "r", encoding="utf-8") as f:
                data = json.load(f)
                obstojece_stevilke = data.get("telefonske_stevilke", [])
        except (json.JSONDecodeError, FileNotFoundError):
            obstojece_stevilke = []

    vse_stevilke = obstojece_stevilke + telefonske_stevilke

    with open(pot_do_datoteke, "w", encoding="utf-8") as f:
        json.dump(
            {"telefonske_stevilke": vse_stevilke},
            f,
            ensure_ascii=False,
            indent=2
        )

    return {
        "message": "Telefonske številke so bile dodane v datoteko.",
        "telefonske_stevilke": telefonske_stevilke,
        "vse_stevilke": vse_stevilke
    }





NAME_HEADERS = {"name", "ime", "full name", "fullname"}
PHONE_HEADERS = {"phone", "telefon", "telefonska", "telefonska stevilka", "telefonska številka", "mobile", "gsm"}


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def normalize_phone(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^\d+]", "", str(value).strip())


def remove_base64_prefix(base64_string: str) -> str:
    if "," in base64_string:
        return base64_string.split(",", 1)[1]
    return base64_string


def obdelaj_xlsx(data: XlsxBase64Request):
    results = []

    for i, file_base64 in enumerate(data.files):
        try:
            clean_base64 = remove_base64_prefix(file_base64)
            file_bytes = base64.b64decode(clean_base64)

            workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)

            file_contacts = []

            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))

                if not rows:
                    continue

                headers = [normalize_header(cell) for cell in rows[0]]

                name_index = None
                phone_index = None

                for j, header in enumerate(headers):
                    if header in NAME_HEADERS and name_index is None:
                        name_index = j
                    if header in PHONE_HEADERS and phone_index is None:
                        phone_index = j

                if name_index is None or phone_index is None:
                    continue

                for row in rows[1:]:
                    name = str(row[name_index]).strip() if name_index < len(row) and row[name_index] is not None else ""
                    phone = normalize_phone(row[phone_index]) if phone_index < len(row) else ""

                    if not name and not phone:
                        continue

                    file_contacts.append({
                        "name": name,
                        "phone": phone
                    })

            results.append({
                "file_name": f"datoteka_{i+1}.xlsx",
                "contacts": file_contacts
            })
        except Exception as e:
            results.append({
                "file_name": f"datoteka_{i+1}.xlsx",
                "error": str(e)
            })

    return results